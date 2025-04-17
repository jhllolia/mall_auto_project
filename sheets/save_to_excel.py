import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, numbers

def save_to_excel(rows, file_path="주문리스트_예시.xlsx"):
    columns = [
        "번호", "구매자명", "상품명", "옵션", "수량",
        "배송지", "연락처", "구매경로", "배송비", "과세", "면세", "배송 메세지"
    ]

    # 빈 줄 제거
    cleaned_rows = [row for row in rows if any(cell != "" for cell in row)]

    # 배송비 무료 표기, 과세/면세 계산 반영 및 숫자 포맷 처리
    for row in cleaned_rows:
        # 배송비
        if isinstance(row[8], (int, float)) and row[8] == 0:
            row[8] = "무료"
        elif isinstance(row[8], (int, float)):
            row[8] = f"{row[8]:,}"

        # 과세 / 면세 * 수량
        try:
            qty = int(row[4])
            tax = int(row[9]) if row[9] not in ("", None) else 0
            free = int(row[10]) if row[10] not in ("", None) else 0
            row[9] = f"{tax * qty:,}" if tax else ""
            row[10] = f"{free * qty:,}" if free else ""
        except:
            pass

    df = pd.DataFrame(cleaned_rows, columns=columns)
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    ws = wb.active

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    thick_black_bottom = Side(style="thick", color="000000")

    previous_order_number = None
    next_order_number = None

    for i in range(2, ws.max_row + 1):
        current_order_number = ws.cell(row=i, column=1).value
        next_order_number = ws.cell(row=i + 1, column=1).value if i + 1 <= ws.max_row else None

        for j in range(1, ws.max_column + 1):
            cell = ws.cell(row=i, column=j)
            col_header = ws.cell(row=1, column=j).value

            # 구매경로 표기 간소화
            if col_header == "구매경로" and cell.value:
                cell.value = "쿠팡" if "쿠팡" in str(cell.value) else "네이버"

            # 번호는 같은 주문에 대해서는 한 번만 출력
            if col_header == "번호":
                if current_order_number == previous_order_number:
                    cell.value = ""
                else:
                    previous_order_number = current_order_number

            # 좌측 정렬 열
            if j in [2, 6, 7, 12] and ws.cell(row=i, column=2).value != "":
                cell.alignment = left
            else:
                cell.alignment = center

            # 기본 테두리
            cell.border = thin_border

            # 주문 마지막 줄에만 하단 굵은 테두리 추가
            if current_order_number != next_order_number:
                cell.border = Border(
                    left=thin_border.left,
                    right=thin_border.right,
                    top=thin_border.top,
                    bottom=thick_black_bottom
                )

    # 열 너비 고정 설정
    col_widths = [6, 12, 28, 22, 6, 35, 15, 10, 10, 10, 10, 25]
    for i, width in enumerate(col_widths, start=1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
        ws.column_dimensions[col_letter].width = width

    wb.save(file_path)
    print(f"✅ 엑셀 저장 완료: {file_path}")
